import openpyxl as px
import os, glob

path = r"D:\D_tokudome\D_desktop\visual_monitoring_file\*"
exel_list = glob.glob(path)
result_book = px.Workbook()
result_book["Sheet"].title = "Sheet1"
result = result_book["Sheet1"]

stock_list = [41, 42, 43, 48]



for n, f in enumerate(exel_list):
    date_name = os.path.splitext(os.path.basename(f))[0]
    wb = px.load_workbook(f)
    sheet = wb["Sheet1"]
    result.cell(row=1, column=n+2).value = date_name
    for m, s in enumerate(stock_list):
        row = s*2 + 1
        row_2 = s*2 + 2
        tsubomi = sheet.cell(row=row, column=2).value
        flower = sheet.cell(row=row_2, column=2).value
        result.cell(row=(m+1)*2, column=1).value = s
        
        result.cell(row=(m+1)*2, column=n+2).value = tsubomi
        result.cell(row=(m+1)*2+1, column=n+2).value = flower
    
result_book.save(r"D:\D_tokudome\D_desktop\data.xlsx")
    