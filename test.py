import openpyxl as px
from openpyxl.styles import PatternFill
import os
import glob
import datetime

#エクセルファイル読み込み
###############エクセルが入っているフォルダのパスに書き換えてください##################
xlfile_folder_path = r"D:\D_tokudome\D_desktop\result_analysis\*.xlsx"
##################################################################################

###############保存先のパスに書き換えてください(tyousa, resultのところが保存ファイル名になっています)###################
#調査表保存用のパス
save_tyousa_path = r"D:\D_tokudome\D_desktop\tyousa_t"
#結果保存用のパス
save_result_path = r"D:\D_tokudome\D_desktop\result_t"
##################################################################################################################

flower_cluster_num = 9

max_value = 0

#ファイル取得
xlfile_path_list = glob.glob(xlfile_folder_path)

#フォルダ生成
os.makedirs(save_tyousa_path, exist_ok=True)
os.makedirs(save_result_path, exist_ok=True)

#色指定
def get_red(plus_red=0):
    hex_red = f"{5+plus_red:02X}{plus_red//4:02X}{plus_red//4:02X}"
    red = PatternFill(patternType="solid", fgColor=hex_red, bgColor=hex_red)
    return red
def get_green(plus_green=0):
    hex_green = f"{5:02X}{5+plus_green:02X}{5:02X}"
    green = PatternFill(patternType="solid", fgColor=hex_green, bgColor=hex_green)
    return green
def get_blue(plus_blue=0):
    hex_blue = f"{0:02X}{0:02X}{150+plus_blue:02X}"
    blue = PatternFill(patternType="solid", fgColor=hex_blue, bgColor=hex_blue)
    return blue

#空のワークブック作成　result書き込み用
wb_uekae_write = px.Workbook()
wb_uekae_write["Sheet"].title = "Sheet1"
sheet_uekae_write = wb_uekae_write["Sheet1"]

wb_uekae = px.Workbook()
wb_uekae["Sheet"].title = "Sheet1"
sheet_uekae = wb_uekae["Sheet1"]

wb_pre = px.Workbook()
wb_pre["Sheet"].title = "Sheet1"
sheet_pre = wb_pre["Sheet1"]

result_book = [0]
tsubomi_book = [0]
flower_book = [0]

result_sheet = [0]
tsubomi_sheet = [0]
flower_sheet = [0]

for i in range(1, flower_cluster_num+1):
    
    result_book.append(px.Workbook())
    result_book[i]["Sheet"].title = "Sheet1"
    result_sheet.append(result_book[i]["Sheet1"])
    
    tsubomi_book.append(px.Workbook())
    tsubomi_book[i]["Sheet"].title = "Sheet1"
    tsubomi_sheet.append(tsubomi_book[i]["Sheet1"])
    
    flower_book.append(px.Workbook())
    flower_book[i]["Sheet"].title = "Sheet1"
    flower_sheet.append(flower_book[i]["Sheet1"])

print(result_book)
    
#出蕾日、開花日記録、色塗り
for n in range(0, len(xlfile_path_list)):
    print(xlfile_path_list[n])
    #現在;;;;;;;のワークブック作成
    wb = px.load_workbook(xlfile_path_list[n])
    sheet_now = wb["Sheet1"]
    
    #sheetの日にち取得
    date_name = os.path.splitext(os.path.basename(xlfile_path_list[n]))[0]
    year = int(date_name[0:4])
    month = int(date_name[4:6])
    day = int(date_name[6:8])
    #取得した日にちに１日足す
    next_date_name = datetime.date(year, month, day) + datetime.timedelta(days=1)
    next_date_name = next_date_name.strftime("%Y%m%d")
    
    #いらない栽培ベッドデータ削除
    for column in range(1, sheet_now.max_column):
        value = sheet_now.cell(row=1, column=column).value
        if value not in (4, 10, 16, 22, 27, None):
            sheet_now.delete_cols(column, 2)
            
        #初日のデータ        
    if n == 0:
        for row in range(1, sheet_now.max_row+1):
            for column in range(1, sheet_now.max_column+1):
                value_init = sheet_now.cell(row=row, column=column).value
                
                #初日のシートをslist[1]にコピー
                sheet_uekae_write.cell(row=row, column=column).value = value_init
                
                
                for i in range(1, flower_cluster_num+1):
                    result_sheet[i].cell(row=row, column=column).value = value_init
                    
                    if 2 >= row:
                        tsubomi_sheet[i].cell(row=row, column=column).value = value_init
                        flower_sheet[i].cell(row=row, column=column).value = value_init
                        sheet_uekae.cell(row=row, column=column).value = value_init
                    #出蕾日、開花日の欄の範囲のとき
                    if (3 <= row) and (2 <= column):
                        if sheet_now.cell(row=row, column=column).value != None:
                            # tsubomi_sheet[i].cell(row=row, column=column).value = 0
                            # flower_sheet[i].cell(row=row, column=column).value = 0
                            if i == 1:    
                                sheet_uekae_write.cell(row=row, column=column).value = 0
                            elif i > 1:
                                result_sheet[i].cell(row=row, column=column).value = 0
                            
                        #1のとき
                        if sheet_now.cell(row=row, column=column).value == 1:
                            #slist[1]の1を日にちに置き換える
                            result_sheet[1].cell(row=row, column=column).value = int(date_name)
                            #各シートの開花日を赤
                            if row % 2 == 1:
                                sheet_now.cell(row=row, column=column).fill = get_green(60)
                            #各シートの出蕾日を緑
                            else:
                                sheet_now.cell(row=row, column=column).fill = get_red(60)
                            
    #２日目以降のデータ                        
    else:
        for row in range(3, sheet_now.max_row+1):
            for column in range(2, sheet_now.max_column+1):
                
                #値がNoneなら前日の値を引き継ぐ
                if row % 2 == 1:
                    if sheet_now.cell(row=row, column=column).value == None:
                        sheet_now.cell(row=row, column=column).value = sheet_pre.cell(row=row, column=column).value
                        sheet_now.cell(row=row+1, column=column).value = sheet_pre.cell(row=row+1, column=column).value
                        
                    #つぼみデータ行と花データ行の値を取得
                    value_tsubomi = sheet_now.cell(row=row, column=column).value
                    value_flower = sheet_now.cell(row=row+1, column=column).value
                    
                    value_pre_tsubomi = sheet_pre.cell(row=row, column=column).value
                    value_pre_flower = sheet_pre.cell(row=row+1, column=column).value
                    
                    #両方０のとき前日のつぼみか花が０でないなら調査票を青に塗る（植え替え）、結果表の日にちを０に戻す       
                    if value_tsubomi == value_flower == 0:
                        
                        if value_pre_tsubomi != 0 or value_pre_flower != 0:
                            #青に塗る
                            sheet_now.cell(row=row, column=column).fill = get_blue()
                            #日にちを０に戻す
                            result_sheet[1].cell(row=row, column=column).value = 0
                            result_sheet[1].cell(row=row+1, column=column).value = 0
                            #植え替え日を記録
                            sheet_uekae_write.cell(row=row, column=column).value = int(date_name)
                        #前日も青なら塗る    
                        if sheet_pre.cell(row=row, column=column).fill == get_blue():
                            sheet_now.cell(row=row, column=column).fill = get_blue()
                            
                    #つぼみデータが０以外の数字のとき
                    if value_tsubomi in range(1, flower_cluster_num+1):
                        #調査票を緑に塗る
                        sheet_now.cell(row=row, column=column).fill = get_green(value_tsubomi*60)
                        #前日つぼみが出ていなければ、結果表に出蕾日を記録
                        if value_pre_tsubomi == 0:
                            result_sheet[value_tsubomi].cell(row=row, column=column).value = int(date_name)
                        if value_tsubomi > max_value:
                            max_value = value_tsubomi

                    #花データが0以外の数字のとき
                    if value_flower in range(1, flower_cluster_num+1):
                        #調査票を赤に塗る
                        sheet_now.cell(row=row+1, column=column).fill = get_red(value_flower*60)
                        #前日花が咲いていなければ、結果表に開花日を記録
                        if value_pre_flower == 0:
                            result_sheet[value_flower].cell(row=row+1, column=column).value = int(date_name)
                        if value_flower > max_value:
                            max_value = value_flower

    for row in range(3, sheet_now.max_row+1):
        for column in range(2, sheet_now.max_column+1):
            sheet_pre.cell(row=row, column=column).value = sheet_now.cell(row=row, column=column).value
            sheet_pre.cell(row=row, column=column).fill = sheet_now.cell(row=row, column=column).fill._StyleProxy__target
    print(max_value) 
            
#３行目からは
for i in range(1, max_value+1):
    for row in range(3, sheet_now.max_row+1):
        for column in range(1, sheet_now.max_column+1):
            value_init = result_sheet[i].cell(row=row, column=column).value
            if i == 1 and column >= 2:
                #調査票の中身を空欄に
                sheet_now.cell(row=row, column=column).value = None
            #tsubomi用にコピー
            if row % 2 == 1:
                tsubomi_sheet[i].cell(row=(row//2)+2, column=column).value = value_init
                if i == 1:
                    sheet_uekae.cell(row=(row//2)+2, column=column).value = sheet_uekae_write.cell(row=row, column=column).value
            #flower用にコピー
            else:
                flower_sheet[i].cell(row=row//2+1, column=column).value = value_init                    
        

#保存_調査票
wb.save(save_tyousa_path + "\\" + str(next_date_name) + ".xlsx") 
for i in range(1, max_value+1):
    result_book[i].save(save_result_path + "\\" + str(date_name) + "_result_" + f"{i:02x}"+ ".xlsx")
    tsubomi_book[i].save(save_result_path + "\\" + str(date_name) + "_result_tsubomi_" + f"{i:02x}"+ ".xlsx")
    flower_book[i].save(save_result_path + "\\" + str(date_name) + "_result_flower_" + f"{i:02x}"+ ".xlsx")
#保存_結果表（出力日は、エクセルリストの最後の日付）
wb_uekae.save(save_result_path + "\\" + str(date_name) + "_result_uekae.xlsx")                          
                            
                            
                            
                            
                            
                            
                            
                            
                            
                            
                            
                            
                            















